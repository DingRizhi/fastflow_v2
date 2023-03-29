from openpyxl import load_workbook


class ParseExcel(object):
    """解析excel文件"""

    def __init__(self, data_path):
        # 获取excel路径
        self.data_path = data_path
        # 获取到excel指定对象
        self.wb = load_workbook(self.data_path)

        """sheet_name是excel最下面的Sheet1、Sheet2、Sheet3....."""

    def get_row_value(self, sheet_name, raw_no):
        """获取某一行的数据"""
        sh = self.wb[sheet_name]
        row_value_list = []
        for y in range(2, sh.max_column + 1):
            value = sh.cell(raw_no, y).value
            row_value_list.append(value)
        return row_value_list

    def get_column_value(self, sheet_name, col_no):
        """获取某一列的数据"""
        sh = self.wb[sheet_name]
        col_value_list = []
        for x in range(2, sh.max_row + 1):
            value = sh.cell(x, col_no).value
            col_value_list.append(value)
        return col_value_list

    def get_cell_value(self, sheet_name, raw_no, col_no):
        """获取某一个单元格的数据"""
        sh = self.wb[sheet_name]
        value = sh.cell(raw_no, col_no).value
        return value

    def write_cell(self, sheet_name, raw_no, col_no, value):
        """向某个单元格写入数据"""
        sh = self.wb[sheet_name]
        sh.cell(raw_no, col_no).value = value
        self.wb.save(self.data_path)